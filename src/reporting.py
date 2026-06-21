import os
import logging
import pandas as pd
import matplotlib.pyplot as plt
from typing import Dict, Any, List

logger = logging.getLogger("QuickCart.Reporting")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenPyXLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl is not installed. Excel report generation will fail.")


class FeedbackReporter:
    """
    Generates Matplotlib chart images and compiles a styled multi-tab Excel Workbook
    as well as a Markdown report for QuickCart Customer Feedback Intelligence.
    """
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Color palette for charts & Excel sheets
        self.colors = {
            'primary': '#1f4e78',      # Dark Navy
            'secondary': '#5b9bd5',    # Slate Blue
            'bg_light': '#f2f4f7',     # Light Grey
            'border': '#d9d9d9',       # Grid Border Grey
            
            # Sentiment soft tones
            'Positive': '#34d399',     # Soft Green
            'Negative': '#fb7185',     # Soft Red
            'Neutral': '#cbd5e1'       # Soft Grey
        }

    def generate_charts(self, df: pd.DataFrame) -> Dict[str, str]:
        """
        Generates feedback category and sentiment distribution plots as PNG files.
        """
        chart_paths = {}
        if df.empty:
            logger.warning("No data loaded. Cannot generate charts.")
            return chart_paths

        # Matplotlib style overrides
        plt.style.use('ggplot')
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Segoe UI', 'DejaVu Sans', 'Arial']

        # 1. Category Distribution Bar Chart
        try:
            fig, ax = plt.subplots(figsize=(6.5, 4))
            cat_counts = df['category'].value_counts().sort_values(ascending=True)
            
            bars = ax.barh(
                cat_counts.index, 
                cat_counts.values, 
                color=self.colors['secondary'],
                edgecolor=self.colors['primary'],
                height=0.6
            )
            
            # Label counts at the end of each bar
            for bar in bars:
                width = bar.get_width()
                ax.text(
                    width + 0.1, 
                    bar.get_y() + bar.get_height()/2, 
                    f'{int(width)}', 
                    va='center', 
                    ha='left', 
                    fontsize=9, 
                    fontweight='bold',
                    color='#333333'
                )

            ax.set_title("Feedback Category Distribution", fontsize=12, fontweight='bold', pad=12, color=self.colors['primary'])
            ax.set_xlabel("Feedback Count", fontsize=9)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_color(self.colors['border'])
            ax.spines['bottom'].set_color(self.colors['border'])
            ax.xaxis.grid(True, linestyle='--', alpha=0.5)
            ax.yaxis.grid(False)
            plt.tight_layout()
            
            cat_chart_path = os.path.join(self.output_dir, "category_distribution.png")
            plt.savefig(cat_chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            chart_paths['category'] = cat_chart_path
            logger.info(f"Category chart saved: {cat_chart_path}")
        except Exception as e:
            logger.error(f"Error compiling category chart: {e}")

        # 2. Sentiment Distribution Donut Chart
        try:
            fig, ax = plt.subplots(figsize=(5.5, 4))
            sent_counts = df['sentiment'].value_counts()
            
            # Map exact colors
            colors_list = [self.colors.get(label, '#cbd5e1') for label in sent_counts.index]
            
            wedges, texts, autotexts = ax.pie(
                sent_counts.values,
                labels=sent_counts.index,
                autopct='%1.1f%%',
                startangle=90,
                colors=colors_list,
                wedgeprops=dict(width=0.4, edgecolor='white', linewidth=2),
                pctdistance=0.75
            )
            
            for t in texts:
                t.set_color('#333333')
                t.set_fontsize(9)
            for at in autotexts:
                at.set_fontsize(9)
                at.set_weight('bold')
                at.set_color('#111111')
                
            ax.set_title("Feedback Sentiment Breakdown", fontsize=12, fontweight='bold', pad=12, color=self.colors['primary'])
            plt.tight_layout()
            
            sent_chart_path = os.path.join(self.output_dir, "sentiment_distribution.png")
            plt.savefig(sent_chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            chart_paths['sentiment'] = sent_chart_path
            logger.info(f"Sentiment chart saved: {sent_chart_path}")
        except Exception as e:
            logger.error(f"Error compiling sentiment chart: {e}")

        return chart_paths

    def generate_excel_report(self, df: pd.DataFrame, chart_paths: Dict[str, str], excel_path: str):
        """
        Compiles a styled multi-tab Excel spreadsheet from the database.
        Embeds the category and sentiment charts directly in the workbook.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("Skipping Excel creation: openpyxl is not installed.")
            return

        wb = openpyxl.Workbook()
        
        # Style Definitions
        font_family = "Segoe UI"
        font_title = Font(name=font_family, size=15, bold=True, color="1F4E78")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="595959")
        font_section = Font(name=font_family, size=11, bold=True, color="1F4E78")
        font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        font_data = Font(name=font_family, size=10)
        font_kpi_label = Font(name=font_family, size=9, bold=True, color="595959")
        font_kpi_val = Font(name=font_family, size=16, bold=True, color="1F4E78")
        
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_kpi = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color="D9D9D9"),
            right=Side(style='thin', color="D9D9D9"),
            top=Side(style='thin', color="D9D9D9"),
            bottom=Side(style='thin', color="D9D9D9")
        )
        border_kpi = Border(
            left=Side(style='medium', color="5B9BD5"),
            right=Side(style='thin', color="D9D9D9"),
            top=Side(style='thin', color="D9D9D9"),
            bottom=Side(style='thin', color="D9D9D9")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # ---------------------------------------------------------
        # TAB 1: Executive Summary
        # ---------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Title
        ws1["A1"] = "QuickCart Customer Feedback Intelligence"
        ws1["A1"].font = font_title
        ws1["A2"] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')} | Dynamic Analysis Output"
        ws1["A2"].font = font_subtitle
        
        # KPI calculations
        total_rows = len(df)
        pos_pct = (df['sentiment'] == 'Positive').mean() if total_rows else 0.0
        neg_pct = (df['sentiment'] == 'Negative').mean() if total_rows else 0.0
        
        neg_df = df[df['sentiment'] == 'Negative']
        top_topic = neg_df['category'].mode().values[0] if not neg_df.empty else "None"

        kpis = [
            ("TOTAL FEEDBACKS", total_rows, "Count"),
            ("POSITIVE RATE", pos_pct, "Percentage"),
            ("NEGATIVE RATE", neg_pct, "Percentage"),
            ("TOP COMPLAINT TOPIC", top_topic, "Text")
        ]
        
        # Write KPIs
        cols = ['A', 'B', 'C', 'D']
        for i, (label, val, val_type) in enumerate(kpis):
            col = cols[i]
            ws1[f"{col}4"] = label
            ws1[f"{col}4"].font = font_kpi_label
            ws1[f"{col}4"].fill = fill_kpi
            ws1[f"{col}4"].alignment = align_center
            ws1[f"{col}4"].border = border_kpi
            
            ws1[f"{col}5"] = val
            ws1[f"{col}5"].font = font_kpi_val
            ws1[f"{col}5"].fill = fill_kpi
            ws1[f"{col}5"].alignment = align_center
            ws1[f"{col}5"].border = border_kpi
            
            if val_type == "Percentage":
                ws1[f"{col}5"].number_format = "0.0%"
            elif val_type == "Count":
                ws1[f"{col}5"].number_format = "#,##0"

        # Category Table (A8:C14)
        ws1["A7"] = "Category Distribution Analysis"
        ws1["A7"].font = font_section
        
        ws1["A8"] = "Category"
        ws1["B8"] = "Count"
        ws1["C8"] = "Percentage"
        for c_let in ['A', 'B', 'C']:
            ws1[f"{c_let}8"].font = font_header
            ws1[f"{c_let}8"].fill = fill_header
            ws1[f"{c_let}8"].alignment = align_center
            ws1[f"{c_let}8"].border = border_thin
            
        cat_counts = df['category'].value_counts()
        row_idx = 9
        for cat, count in cat_counts.items():
            ws1[f"A{row_idx}"] = cat
            ws1[f"A{row_idx}"].font = font_data
            ws1[f"A{row_idx}"].alignment = align_left
            ws1[f"A{row_idx}"].border = border_thin
            
            ws1[f"B{row_idx}"] = count
            ws1[f"B{row_idx}"].font = font_data
            ws1[f"B{row_idx}"].alignment = align_right
            ws1[f"B{row_idx}"].number_format = "#,##0"
            ws1[f"B{row_idx}"].border = border_thin
            
            ws1[f"C{row_idx}"] = count / total_rows if total_rows else 0
            ws1[f"C{row_idx}"].font = font_data
            ws1[f"C{row_idx}"].alignment = align_right
            ws1[f"C{row_idx}"].number_format = "0.0%"
            ws1[f"C{row_idx}"].border = border_thin
            row_idx += 1

        # Sentiment Table (A16:C20)
        ws1["A15"] = "Sentiment Distribution Analysis"
        ws1["A15"].font = font_section
        
        ws1["A16"] = "Sentiment"
        ws1["B16"] = "Count"
        ws1["C16"] = "Percentage"
        for c_let in ['A', 'B', 'C']:
            ws1[f"{c_let}16"].font = font_header
            ws1[f"{c_let}16"].fill = fill_header
            ws1[f"{c_let}16"].alignment = align_center
            ws1[f"{c_let}16"].border = border_thin
            
        sent_counts = df['sentiment'].value_counts()
        row_idx = 17
        for sent in ['Positive', 'Negative', 'Neutral']:
            count = sent_counts.get(sent, 0)
            ws1[f"A{row_idx}"] = sent
            ws1[f"A{row_idx}"].font = font_data
            ws1[f"A{row_idx}"].alignment = align_left
            ws1[f"A{row_idx}"].border = border_thin
            
            ws1[f"B{row_idx}"] = count
            ws1[f"B{row_idx}"].font = font_data
            ws1[f"B{row_idx}"].alignment = align_right
            ws1[f"B{row_idx}"].number_format = "#,##0"
            ws1[f"B{row_idx}"].border = border_thin
            
            ws1[f"C{row_idx}"] = count / total_rows if total_rows else 0
            ws1[f"C{row_idx}"].font = font_data
            ws1[f"C{row_idx}"].alignment = align_right
            ws1[f"C{row_idx}"].number_format = "0.0%"
            ws1[f"C{row_idx}"].border = border_thin
            row_idx += 1

        # Add Charts
        if 'category' in chart_paths and os.path.exists(chart_paths['category']):
            img_cat = OpenPyXLImage(chart_paths['category'])
            img_cat.width, img_cat.height = 420, 260
            ws1.add_image(img_cat, 'E3')
            
        if 'sentiment' in chart_paths and os.path.exists(chart_paths['sentiment']):
            img_sent = OpenPyXLImage(chart_paths['sentiment'])
            img_sent.width, img_sent.height = 360, 260
            ws1.add_image(img_sent, 'E17')

        # Adjust columns
        for col_let in ['A', 'B', 'C', 'D']:
            ws1.column_dimensions[col_let].width = 22
        ws1.column_dimensions['E'].width = 32

        # ---------------------------------------------------------
        # TAB 2: Representative Examples
        # ---------------------------------------------------------
        ws2 = wb.create_sheet(title="Representative Examples")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "Key Customer Complaints & Representative Examples"
        ws2["A1"].font = font_title
        ws2["A2"] = "Curated customer voices mapped to categories (prioritizing negative reviews)"
        ws2["A2"].font = font_subtitle
        
        headers2 = ["Category", "Rating", "Source", "Raw Feedback Text", "AI Issue Summary"]
        cols2 = ['A', 'B', 'C', 'D', 'E']
        
        ws2.row_dimensions[4].height = 24
        for idx, h in enumerate(headers2):
            c_let = cols2[idx]
            ws2[f"{c_let}4"] = h
            ws2[f"{c_let}4"].font = font_header
            ws2[f"{c_let}4"].fill = fill_header
            ws2[f"{c_let}4"].alignment = align_center
            ws2[f"{c_let}4"].border = border_thin

        # Pull samples per category
        row_idx = 5
        for cat in df['category'].unique():
            cat_df = df[df['category'] == cat]
            
            # Prioritize negative feedback for complaint examples
            neg_samples = cat_df[cat_df['sentiment'] == 'Negative']
            samples = neg_samples if not neg_samples.empty else cat_df
            samples = samples.head(3)
            
            for _, r in samples.iterrows():
                ws2.row_dimensions[row_idx].height = 40
                
                ws2[f"A{row_idx}"] = r['category']
                ws2[f"A{row_idx}"].alignment = align_center
                
                ws2[f"B{row_idx}"] = r['rating'] if pd.notnull(r['rating']) else "N/A"
                ws2[f"B{row_idx}"].alignment = align_center
                
                ws2[f"C{row_idx}"] = str(r['source']).replace('_', ' ').title()
                ws2[f"C{row_idx}"].alignment = align_center
                
                ws2[f"D{row_idx}"] = r['feedback_text']
                ws2[f"D{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                ws2[f"E{row_idx}"] = r['issue_summary']
                ws2[f"E{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                for c_let in cols2:
                    ws2[f"{c_let}{row_idx}"].font = font_data
                    ws2[f"{c_let}{row_idx}"].border = border_thin
                    if row_idx % 2 == 0:
                        ws2[f"{c_let}{row_idx}"].fill = fill_zebra
                row_idx += 1

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 50
        ws2.column_dimensions['E'].width = 40

        # ---------------------------------------------------------
        # TAB 3: All Feedback Data
        # ---------------------------------------------------------
        ws3 = wb.create_sheet(title="All Feedback Data")
        ws3.views.sheetView[0].showGridLines = True
        
        ws3["A1"] = "Complete Enriched Customer Feedback"
        ws3["A1"].font = font_title
        ws3["A2"] = "Cleaned customer feedbacks enriched with sentiment, category and summary"
        ws3["A2"].font = font_subtitle
        
        headers3 = ["ID", "Timestamp", "Source", "Rating", "Feedback Text", "Sentiment", "Category", "AI Issue Summary"]
        cols3 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        
        ws3.row_dimensions[4].height = 24
        for idx, h in enumerate(headers3):
            c_let = cols3[idx]
            ws3[f"{c_let}4"] = h
            ws3[f"{c_let}4"].font = font_header
            ws3[f"{c_let}4"].fill = fill_header
            ws3[f"{c_let}4"].alignment = align_center
            ws3[f"{c_let}4"].border = border_thin

        row_idx = 5
        for _, r in df.iterrows():
            ws3.row_dimensions[row_idx].height = 24
            
            ws3[f"A{row_idx}"] = r['id']
            ws3[f"A{row_idx}"].alignment = align_center
            
            ws3[f"B{row_idx}"] = r['timestamp'] if pd.notnull(r['timestamp']) else "N/A"
            ws3[f"B{row_idx}"].alignment = align_center
            
            ws3[f"C{row_idx}"] = str(r['source']).replace('_', ' ').title()
            ws3[f"C{row_idx}"].alignment = align_center
            
            ws3[f"D{row_idx}"] = r['rating'] if pd.notnull(r['rating']) else "N/A"
            ws3[f"D{row_idx}"].alignment = align_center
            
            ws3[f"E{row_idx}"] = r['feedback_text']
            ws3[f"E{row_idx}"].alignment = align_left
            
            ws3[f"F{row_idx}"] = r['sentiment']
            ws3[f"F{row_idx}"].alignment = align_center
            
            # Highlight sentiment background softly
            if r['sentiment'] == 'Positive':
                ws3[f"F{row_idx}"].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif r['sentiment'] == 'Negative':
                ws3[f"F{row_idx}"].fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
            else:
                ws3[f"F{row_idx}"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                
            ws3[f"G{row_idx}"] = r['category']
            ws3[f"G{row_idx}"].alignment = align_center
            
            ws3[f"H{row_idx}"] = r['issue_summary']
            ws3[f"H{row_idx}"].alignment = align_left
            
            for c_let in cols3:
                ws3[f"{c_let}{row_idx}"].font = font_data
                ws3[f"{c_let}{row_idx}"].border = border_thin
                if c_let != 'F' and row_idx % 2 == 0:
                    ws3[f"{c_let}{row_idx}"].fill = fill_zebra
                    
            row_idx += 1

        # Adjust columns dynamically
        for c_let in cols3:
            max_len = 0
            for cell_row in range(4, row_idx):
                val = ws3[f"{c_let}{cell_row}"].value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws3.column_dimensions[c_let].width = min(max(max_len + 4, 12), 60)

        # Save workbook
        wb.save(excel_path)
        logger.info(f"Excel report saved successfully to: {excel_path}")

    def generate_markdown_report(self, df: pd.DataFrame, report_path: str):
        """
        Produces a decision-ready markdown summary report for leadership.
        """
        if df.empty:
            logger.warning("Empty DataFrame. Cannot compile markdown report.")
            return

        total = len(df)
        cat_counts = df['category'].value_counts()
        sent_counts = df['sentiment'].value_counts()
        top5_cats = cat_counts.head(5)

        lines = [
            "# QuickCart Customer Feedback Intelligence Summary",
            "",
            f"**Generated:** {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}  ",
            f"**Total Customer Feedbacks Analyzed:** {total:,}",
            "",
            "---",
            "",
            "## Top 5 Complaint Categories by Volume",
            "",
            "| Rank | Category | Count | % of Total |",
            "|------|----------|------:|-----------:|",
        ]

        for rank, (cat, count) in enumerate(top5_cats.items(), start=1):
            pct = count / total * 100
            lines.append(f"| {rank} | {cat} | {count:,} | {pct:.1f}% |")

        lines.extend([
            "",
            "## Overall Sentiment Breakdown",
            "",
            "| Sentiment | Count | Percentage |",
            "|-----------|------:|-----------:|",
        ])

        for sent in ['Positive', 'Negative', 'Neutral']:
            count = sent_counts.get(sent, 0)
            pct = count / total * 100 if total else 0
            lines.append(f"| {sent} | {count:,} | {pct:.1f}% |")

        neg_pct = sent_counts.get('Negative', 0) / total * 100 if total else 0
        pos_pct = sent_counts.get('Positive', 0) / total * 100 if total else 0

        lines.extend([
            "",
            f"**Key Sentiment Ratio:** Customer experience is **{neg_pct:.1f}% Negative** vs **{pos_pct:.1f}% Positive**.",
            "",
            "---",
            "",
            "## Representative Customer Voices (Examples)",
            "",
        ])

        for cat in top5_cats.index:
            cat_df = df[df['category'] == cat]
            neg_examples = cat_df[cat_df['sentiment'] == 'Negative']
            examples = neg_examples if not neg_examples.empty else cat_df
            examples = examples.head(3)

            lines.append(f"### {cat} ({cat_counts[cat]:,} feedbacks)")
            lines.append("")
            for i, (_, row) in enumerate(examples.iterrows(), start=1):
                rating = int(row['rating']) if pd.notnull(row['rating']) else 'N/A'
                lines.append(f"**Example {i}** *(source: {row['source'].replace('_', ' ').title()}, rating: {rating})*")
                lines.append(f"> \"{row['feedback_text']}\"")
                lines.append(f"> *AI issue summary: {row['issue_summary']}*")
                lines.append("")

        # Trend analysis
        lines.extend(["---", "", "## Trend: Is It Getting Better or Worse?", ""])
        ts_df = df.copy()
        ts_df['month'] = pd.to_datetime(ts_df['timestamp'], errors='coerce').dt.to_period('M')
        dated = ts_df.dropna(subset=['month'])

        if len(dated) >= 2:
            monthly = (
                dated.groupby('month')
                .agg(
                    total=('id', 'count'),
                    negative_pct=('sentiment', lambda s: (s == 'Negative').mean() * 100)
                )
                .sort_index()
            )
            first = monthly.iloc[0]
            last = monthly.iloc[-1]
            delta = last['negative_pct'] - first['negative_pct']
            direction = "worsening" if delta > 2 else ("improving" if delta < -2 else "stable")

            lines.append(
                f"Comparing early dataset records from **{first.name}** ({first['negative_pct']:.1f}% negative) "
                f"to recent **{last.name}** ({last['negative_pct']:.1f}% negative): "
                f"customer experience appears **{direction}** "
                f"({'+' if delta >= 0 else ''}{delta:.1f} percentage points)."
            )
            lines.append("")
            lines.append("| Month | Total Feedbacks | % Negative Sentiment |")
            lines.append("|-------|----------------:|---------------------:|")
            for month, row in monthly.iterrows():
                lines.append(f"| {month} | {int(row['total']):,} | {row['negative_pct']:.1f}% |")
        else:
            lines.append(
                "Insufficient dated records to compute a reliable month-over-month trend. "
                "Many raw timestamps were blank, unparseable, or fell within a single month."
            )

        lines.extend([
            "",
            "---",
            "",
            "*Report compiled automatically by the QuickCart Feedback Intelligence pipeline.*",
        ])

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        logger.info(f"Markdown summary report saved to: {report_path}")
